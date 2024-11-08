from flask import Flask, request, send_file, render_template
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy
from openpyxl.utils import get_column_letter
import os
import pandas as pd
from werkzeug.utils import secure_filename
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

app = Flask(__name__)

# Configure upload folder
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_excel():
    if 'file' not in request.files:
        return 'No file uploaded'
    
    file = request.files['file']
    if file.filename == '':
        return 'No file selected'
        
    if not file.filename.endswith('.xlsx'):
        return 'Please upload an Excel file (.xlsx)'
    
    # Get processor type from form data right after file validation
    processor_type = request.form.get('processor-type', '')
    
    # Save uploaded file
    filename = secure_filename(file.filename)
    input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'processed_{filename}')
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.pdf')
    file.save(input_path)
    
    # Read Excel file with pandas, specifically from 'Data' sheet
    try:
        df = pd.read_excel(input_path, sheet_name='Data')
        print("\nFinal data structure:")
        print(df.head())
        print("\nColumns:", df.columns.tolist())
        
    except Exception as e:
        return f'Error processing file: {str(e)}'
    
    print("\nBefore sorting - Columns:", df.columns.tolist())
    
    # Remove rows containing 'Buffer' in TO_LOC column
    df = df[~df['TO_LOC'].str.contains('Buffer', na=False)]
    
    # Filter rows to only keep those with HFB values 14 or 15
    df = df[df['HFB'].isin([14, 15])]
    
    # Sort data by SLID_P column
    df = df.sort_values('SLID_P')

    # Pad zzzzzz column values with leading zeros to length 8
    if 'ARTNO' in df.columns:
        df['ARTNO'] = df['ARTNO'].astype(str).apply(lambda x: x.zfill(8))
    
    # Save filtered data to new workbook
    workbook = load_workbook(input_path)
    sheet = workbook['Data']  # Specifically use the Data sheet
    
    # Clear existing data (except header)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Write filtered data back to sheet
    for _, row in df.iterrows():
        sheet.append(row.tolist())
    
    # Hide specified columns instead of removing them
    columns_to_hide = ['C', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'V', 'X', 'Y', 'Z', 'AA', 'AB']
    for col in columns_to_hide:
        col_letter = col
        sheet.column_dimensions[col_letter].hidden = True
    # Add asterisks to ARTNO column values
    artno_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'ARTNO':
            artno_col = col
            break
            
    if artno_col:
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
            cell = sheet.cell(row=row, column=artno_col)
            if cell.value:  # Only modify if cell has a value
                cell.value = f'*{str(cell.value)}*'
                
    # Change the font and size
    barcode_font = Font(name="Libre Barcode 39 Text", size=40)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply center alignment to all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
    
    # Apply special font to column A while preserving center alignment
    for cell in sheet['A']:
        cell.font = barcode_font
        
    # Set column A width to 120 pixels
    sheet.column_dimensions['A'].width = 120/7  # Convert pixels to Excel width units (1 unit ≈ 7 pixels)
        
    # Autofit column widths based on content (except column A)
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Skip hidden columns and column A
        if column_letter in columns_to_hide or column_letter == 'A':
            continue
            
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Add a little extra width for padding
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add right border to DEL_TYPE column
    thin_border = Border(right=Side(style='thin'))
    del_type_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'DEL_TYPE':
            del_type_col = col
            break
    
    if del_type_col:
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=del_type_col).border = thin_border
    # Create a new sheet named Data2
    data2 = workbook.create_sheet("Data2")
    
    # Define the columns to copy
    columns_to_copy = ['ARTNO', 'ARTNAME', 'HFB', 'PA', 'SLID_P', 'SLID_H', 'TO_LOC', 'MOVED_QTY', 'DEL_TYPE']
    
    # Find column indices in source sheet
    source_cols = []
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value in columns_to_copy:
            source_cols.append(col)
    
    # Copy values and formatting
    for new_col, src_col in enumerate(source_cols, start=1):
        # Copy column header
        src_header = sheet.cell(row=1, column=src_col)
        dest_header = data2.cell(row=1, column=new_col)
        dest_header.value = src_header.value
        
        # Set Calibri font for "zzzzzz" header
        if dest_header.value == 'ARTNO':
            dest_header.font = Font(name='Calibri')
        else:
            dest_header.font = copy(src_header.font)
            
        dest_header.alignment = copy(src_header.alignment)
        
        # Add gridlines to header
        dest_header.border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))
        
        # Copy column values and formatting
        for row in range(2, sheet.max_row + 1):
            src_cell = sheet.cell(row=row, column=src_col)
            dest_cell = data2.cell(row=row, column=new_col)
            dest_cell.value = src_cell.value
            dest_cell.font = copy(src_cell.font)
            dest_cell.alignment = copy(src_cell.alignment)
            
            # Add gridlines to all cells with values
            if dest_cell.value is not None:
                dest_cell.border = Border(left=Side(style='thin'),
                                       right=Side(style='thin'),
                                       top=Side(style='thin'),
                                       bottom=Side(style='thin'))
        
        # Copy column dimensions
        src_letter = get_column_letter(src_col)
        dest_letter = get_column_letter(new_col)
        data2.column_dimensions[dest_letter].width = sheet.column_dimensions[src_letter].width
        data2.column_dimensions[dest_letter].hidden = sheet.column_dimensions[src_letter].hidden
    # Set page layout to landscape with narrow margins
    data2.page_setup.orientation = data2.ORIENTATION_LANDSCAPE
    data2.page_margins.left = 0.25
    data2.page_margins.right = 0.25
    data2.page_margins.top = 0.25
    data2.page_margins.bottom = 0.25
    data2.page_margins.header = 0.25
    data2.page_margins.footer = 0.25
    # Save the modified workbook
    workbook.save(output_path)
    
    # Register Libre Barcode 128 Text font for PDF
    font_path = './LibreBarcode39Text-Regular.ttf'
    pdfmetrics.registerFont(TTFont('LibreBarcode39Text', font_path))
    
    # Convert to PDF using reportlab
    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter))
    elements = []
    
    # Get data from Data2 sheet
    data = []
    for row in data2.rows:
        data.append([cell.value for cell in row])
    
    # Create table
    # Sort data based on HFB, PA and SLID_H columns
    headers = data[0]
    hfb_idx = headers.index('HFB')
    pa_idx = headers.index('PA')
    slid_h_idx = headers.index('SLID_H')
    
    sorted_data = [data[0]]  # Keep headers as first row
    sorted_data.extend(sorted(data[1:], key=lambda x: (str(x[hfb_idx]), str(x[pa_idx]), str(x[slid_h_idx]))))
    
    table = Table(sorted_data, rowHeights=[40] * len(sorted_data))  # Set uniform row height of 40 points
    table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        # Apply Libre Barcode font to first column (ARTNO)
        ('FONTNAME', (0, 1), (0, -1), 'LibreBarcode39Text'),
        ('FONTSIZE', (0, 1), (0, -1), 22),
        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),  # Vertically align first cell to middle
        ('VALIGN', (0, 1), (0, -1), 'TOP'),  # Vertically align rest of ARTNO column to top
        ('VALIGN', (1, 0), (-1, -1), 'MIDDLE'),  # Center align all other cells vertically
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    # Return the PDF file
    return send_file(pdf_path, as_attachment=True, download_name='processed_report.pdf')

if __name__ == '__main__':
    app.run(debug=True)
